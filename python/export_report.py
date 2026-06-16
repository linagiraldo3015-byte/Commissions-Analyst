# ============================================================
# 03_export_report.py
# Commissions Liquidation System — Bold Fintech (simulated)
# Reads liquidation results from PostgreSQL and exports
# a formatted Excel report ready for Payroll
# ============================================================

import pandas as pd
from sqlalchemy import create_engine
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 1. Connect to PostgreSQL ─────────────────────────────────
engine = create_engine('postgresql://postgres:postgres123@localhost:5432/commissions_db')

# ── 2. Run liquidation query ─────────────────────────────────
query = """
WITH clean_sales AS (
    SELECT
        agent_id,
        COUNT(*)                   AS total_activations,
        SUM(transaction_value_usd) AS total_value_usd
    FROM commissions.sales
    WHERE status = 'Activo'
      AND is_duplicate = FALSE
      AND transaction_value_usd IS NOT NULL
      AND sale_date BETWEEN '2024-01-01' AND '2024-01-31'
    GROUP BY agent_id
),
tiered_sales AS (
    SELECT
        cs.agent_id,
        a.agent_name,
        a.channel,
        a.region,
        a.monthly_quota,
        cs.total_activations,
        cs.total_value_usd,
        sch.tier,
        sch.commission_pct,
        sch.bonus_pct
    FROM clean_sales cs
    JOIN commissions.agents a ON cs.agent_id = a.agent_id
    JOIN commissions.commission_schemes sch
        ON a.channel = sch.channel
        AND cs.total_activations >= sch.min_activations
        AND (cs.total_activations <= sch.max_activations OR sch.max_activations IS NULL)
),
final_report AS (
    SELECT
        agent_id,
        agent_name,
        channel,
        region,
        monthly_quota,
        total_activations,
        ROUND(total_value_usd::NUMERIC, 2)                    AS total_value_usd,
        tier,
        ROUND((total_value_usd * commission_pct)::NUMERIC, 2) AS base_commission_usd,
        CASE
            WHEN total_activations >= monthly_quota
            THEN ROUND((total_value_usd * bonus_pct)::NUMERIC, 2)
            ELSE 0
        END                                                    AS bonus_usd,
        ROUND((total_value_usd * commission_pct +
            CASE
                WHEN total_activations >= monthly_quota
                THEN total_value_usd * bonus_pct
                ELSE 0
            END)::NUMERIC, 2)                                  AS total_payout_usd,
        CASE
            WHEN total_activations >= monthly_quota
            THEN 'Cumplió meta'
            ELSE 'No cumplió meta'
        END                                                    AS quota_status
    FROM tiered_sales
)
SELECT * FROM final_report
ORDER BY total_payout_usd DESC;
"""

df = pd.read_sql(query, engine)
print(f"✅ Query ejecutada: {len(df)} agentes en el reporte")

# ── 3. Rename columns for Payroll ────────────────────────────
df.columns = [
    'ID Agente', 'Nombre', 'Canal', 'Región', 'Meta Mensual',
    'Activaciones', 'Valor Transaccionado (USD)', 'Tier',
    'Comisión Base (USD)', 'Bono (USD)', 'Total a Pagar (USD)', 'Estado Meta'
]

# ── 4. Export to Excel with formatting ───────────────────────
output_path = 'output/liquidacion_enero_2024.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Liquidación Enero 2024', index=False)

    workbook  = writer.book
    worksheet = writer.sheets['Liquidación Enero 2024']

    # Header style
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col_num, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row colors — alternating + highlight quota status
    green_fill  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    red_fill    = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    alt_fill    = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    for row_num in range(2, len(df) + 2):
        quota_status = worksheet.cell(row=row_num, column=12).value
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            if quota_status == 'Cumplió meta':
                cell.fill = green_fill
            elif quota_status == 'No cumplió meta':
                cell.fill = red_fill
            elif row_num % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = [10, 25, 12, 15, 14, 14, 26, 10, 22, 14, 22, 18]
    for i, width in enumerate(col_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width

    # Summary sheet
    summary_data = {
        'Métrica': [
            'Total Agentes',
            'Agentes que Cumplieron Meta',
            '% Cumplimiento',
            'Promedio Activaciones',
            'Costo Total Comisiones (USD)'
        ],
        'Valor': [
            len(df),
            len(df[df['Estado Meta'] == 'Cumplió meta']),
            f"{len(df[df['Estado Meta'] == 'Cumplió meta']) / len(df) * 100:.1f}%",
            f"{df['Activaciones'].mean():.1f}",
            f"${df['Total a Pagar (USD)'].sum():,.2f}"
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)

    ws2 = writer.sheets['Resumen Ejecutivo']
    for col_num in range(1, 3):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 25

print(f"✅ Excel exportado: {output_path}")
print(f"\n📊 Resumen Ejecutivo:")
print(f"   Total agentes:        {len(df)}")
print(f"   Cumplieron meta:      {len(df[df['Estado Meta'] == 'Cumplió meta'])}")
print(f"   Costo total:          ${df['Total a Pagar (USD)'].sum():,.2f} USD")